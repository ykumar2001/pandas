from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Load the existing workbook and select the active sheet
file_path = r"C:\Users\yk687_000\Downloads\excelsheet_format.xlsx"
wb = load_workbook(file_path)
ws = wb.active

# Add a new row for percentages
total_rows = ws.max_row
total_cols = ws.max_column

# Write the "Filled Percentage" label in the first column of a new row
percent_row = total_rows + 2  # Add two rows below existing data (one for spacing)
ws.cell(row=percent_row, column=1, value="Filled Percentage")

# Add formulas for percentage calculation
for col in range(2, total_cols + 1):  # Skip the first column (assume it's headers/labels)
    # Column letter
    col_letter = get_column_letter(col)

    # Formula: (COUNTA(Column) / Total Rows) * 100
    formula = f"=COUNTA({col_letter}2:{col_letter}{total_rows}) / {total_rows} * 100"

    # Write the formula in the percentage row
    ws.cell(row=percent_row, column=col, value=formula)

# Save the workbook
output_path = r"C:\Users\yk687_000\Downloads\excelsheet_with_percentages.xlsx"
wb.save(output_path)

print(f"Percentage formulas added and saved to {output_path}")
