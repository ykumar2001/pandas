from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

file=r"C:\Users\yk687_000\Downloads\excelsheet_format.xlsx"

wb=load_workbook(file)
ws=wb.active

total_rows=ws.max_row
total_col=ws.max_column

ws.cell(row=total_rows+2,column=1,value="total percentage")

for col in range(2,total_col+1):
    col_letter=get_column_letter(col)
    formula=f"=COUNTA({col_letter}2:{col_letter}{total_rows})/{total_rows}*100"
    ws.cell(row=total_rows+2,column=col,value=formula)

outer_path=r"C:\Users\yk687_000\Downloads\percentages.xlsx"
wb.save(outer_path)
print(f"saved your file successfully:{outer_path}")
